import streamlit as st
# [최적화] 무거운 라이브러리(pandas, openpyxl, smtplib 등)의 상단 import를 제거하고
# 필요한 함수 내부에서만 import 하여 앱 초기 실행 속도를 높였습니다.

# ==========================================
# [사용자 설정] 수정 가능한 영역
# ==========================================

# 1. 메일 제목 앞부분 고정 문구
# 결과 예시: "[한국투자증권] 금리데이터송부 2026-02-11" (날짜 앞에 띄어쓰기 포함됨)
FIXED_TITLE_SUFFIX = "[한국투자증권] 금리데이터송부" 

# 2. 메일 본문 하단 서명
FIXED_FOOTER = """
감사합니다.

한국투자증권 FICC Sales부

송인호 부서장 02-3276-5318
권서연 대  리 02-3276-5472
문찬호 대  리 02-3276-6496
김종선 대  리 02-3276-5472
진태영 사  원 02-3276-4976

부서 Email: A07910@koreainvestment.com
"""
# ==========================================

# 1. 페이지 설정
st.set_page_config(page_title="Rates", layout="centered")

# 2. 엑셀 데이터 읽기 (캐싱 + 메모리 최적화)
@st.cache_data(max_entries=1, show_spinner=False)
def get_rates(uploaded_file):
    v_3m, v_3y, v_10y = "", "", ""
    try:
        filename = uploaded_file.name.lower()
        
        # .xls 파일 처리 (구버전)
        if filename.endswith('.xls'):
            import xlrd # [최적화] 필요할 때만 로드
            file_data = uploaded_file.read()
            wb = xlrd.open_workbook(file_contents=file_data)
            sheet = wb.sheet_by_index(0)
            v_3m = str(sheet.cell_value(1, 4))
            v_3y = str(sheet.cell_value(1, 11))
            v_10y = str(sheet.cell_value(1, 15))
            # 메모리 해제
            del file_data, wb, sheet
            
        # .xlsx 파일 처리 (신버전)
        else:
            import openpyxl # [최적화] 필요할 때만 로드
            # read_only=True로 메모리 사용량 최소화
            wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True)
            sheet = wb.active
            v_3m = str(sheet.cell(row=2, column=5).value)
            v_3y = str(sheet.cell(row=2, column=12).value)
            v_10y = str(sheet.cell(row=2, column=16).value)
            wb.close()
            del wb, sheet
            
    except Exception:
        pass
    
    # None값 처리 헬퍼 함수
    def clean(v): return v if v and v != "None" else ""
    return clean(v_3m), clean(v_3y), clean(v_10y)

# 3. 설정 로드 (st.secrets)
secrets = st.secrets.get("gmail", {})
sid = secrets.get("id", "")
spw = secrets.get("pw", "")
srcv = secrets.get("receiver", "")

# 4. UI 구성
uploaded_file = st.file_uploader("Excel", type=["xls", "xlsx"], label_visibility="collapsed")

# 파일 업로드 시 데이터 추출
if uploaded_file:
    v_3m, v_3y, v_10y = get_rates(uploaded_file)
else:
    v_3m, v_3y, v_10y = "", "", ""

with st.form("main"):
    rcv = st.text_input("To", value=srcv, placeholder="Receiver")
    
    cd = st.text_input("CD (%)", placeholder="Enter CD")
    c1, c2, c3 = st.columns(3)
    k3m = c1.text_input("3M", value=v_3m)
    k3y = c2.text_input("3Y", value=v_3y)
    k10y = c3.text_input("10Y", value=v_10y)
    
    attach_excel = st.checkbox("Attach Excel File", value=True)
    
    # 버튼을 꽉 차게 설정하여 모바일 편의성 증대
    submitted = st.form_submit_button("Send", type="primary", use_container_width=True)

# 5. 전송 로직 (실행 속도 최적화 영역)
if submitted:
    if not (sid and spw and rcv):
        st.error("Check Secrets")
    elif not (cd and k3m and k3y and k10y):
        st.warning("Input Data")
    else:
        try:
            # [최적화] 전송 버튼을 눌렀을 때만 무거운 라이브러리를 로드함
            import smtplib
            import io
            import openpyxl
            from email.message import EmailMessage
            from datetime import datetime
            
            # 날짜 계산
            now = datetime.now()
            today_str = now.strftime('%Y-%m-%d')
            file_date_str = now.strftime('%Y%m%d')
            
            msg = EmailMessage()
            
            # [제목] 고정문구 + 띄어쓰기 + 날짜
            msg['Subject'] = f"{FIXED_TITLE_SUFFIX} {today_str}"
            
            msg['From'] = sid
            msg['To'] = rcv
            
            # [본문] 금리 데이터 + 하단 서명
            body_txt = f"""Market Rates ({today_str})

CD: {cd}%
3M: {k3m}%
3Y: {k3y}%
10Y: {k10y}%

---
{FIXED_FOOTER}"""
            msg.set_content(body_txt)

            # [첨부파일 생성]
            if attach_excel:
                wb_new = openpyxl.Workbook()
                ws_new = wb_new.active
                ws_new.title = "Sheet1"
                
                # 숫자 변환
                try:
                    val_cd = float(cd)
                    val_3m = float(k3m)
                    val_3y = float(k3y)
                    val_10y = float(k10y)
                except ValueError:
                    val_cd, val_3m, val_3y, val_10y = cd, k3m, k3y, k10y

                # 데이터 리스트 (세로형, 1행부터 시작)
                data_rows = [
                    [file_date_str, "CD수익률", val_cd],
                    [file_date_str, "KTB3m", val_3m],
                    [file_date_str, "KTB3y", val_3y],
                    [file_date_str, "KTB10y", val_10y]
                ]
                
                for row in data_rows:
                    ws_new.append(row)

                # 서식 지정 (1~4행의 C열, 소수점 2자리)
                for r in range(1, 5): 
                    cell = ws_new.cell(row=r, column=3)
                    cell.number_format = '0.00'

                # 메모리에 엑셀 저장 (디스크 I/O 없음)
                excel_buffer = io.BytesIO()
                wb_new.save(excel_buffer)
                excel_buffer.seek(0)
                file_data = excel_buffer.read()
                
                file_name = f"금리data_{file_date_str}.xlsx"
                
                msg.add_attachment(
                    file_data,
                    maintype='application',
                    subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    filename=file_name
                )
                
                # 사용 후 메모리 정리
                del wb_new, excel_buffer

            # SMTP 전송
            with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                smtp.login(sid, spw)
                smtp.send_message(msg)
            
            st.success(f"Sent! ({file_name})" if attach_excel else "Sent!")
            
        except Exception as e:
            st.error(f"Error: {e}")
